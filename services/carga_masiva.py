import json
import re
import unicodedata
import uuid
from collections import defaultdict
from numbers import Number
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from database.modelos import Comuna, Envio
from utils.texto import normalizar_nombre_operativo, normalizar_texto_operativo
from utils.validaciones import (
    email_valido,
    normalizar_telefono_chile,
    rut_operativo_valido,
    telefono_chile_valido,
)


COLUMNAS_PLANTILLA_ANTERIOR = [
    "Remitente",
    "Correo remitente",
    "Centro costo",
    "Division",
    "Destinatario",
    "RUT destinatario",
    "Direccion",
    "Region",
    "Comuna",
    "Telefono",
    "Correo destinatario",
    "Tipo envio",
    "Bultos",
    "Kilos",
    "Observacion",
]

COLUMNAS_ENVIOS_PLANTILLA = [
    "Destinatario",
    "RUT destinatario",
    "Direccion",
    "Comuna",
    "Region",
    "Telefono",
    "Correo destinatario",
    "Tipo envio",
    "Bultos",
    "Kilos",
    "Observacion",
]

DIVISIONES = ["DPGP", "DOP", "LDB", "DL", "DPP"]
TIPOS_ENVIO = ["Domicilio", "Agencia"]
MAX_FILAS_CARGA = 300
TMP_CARGAS_DIR = Path(__file__).resolve().parent.parent / "tmp_cargas"


MAPA_COLUMNAS = {
    "remitente": "remitente",
    "correo remitente": "correo_remitente",
    "centro costo": "centro_costo",
    "division": "division",
    "destinatario": "destinatario",
    "rut destinatario": "rut_destinatario",
    "rut": "rut_destinatario",
    "direccion": "direccion",
    "region": "region",
    "comuna": "comuna",
    "telefono": "telefono_destinatario",
    "numero telefono": "telefono_destinatario",
    "correo destinatario": "correo_destinatario",
    "e-mail destinatario": "correo_destinatario",
    "email destinatario": "correo_destinatario",
    "e-mail desti": "correo_destinatario",
    "tipo envio": "tipo_envio",
    "tipo de envio": "tipo_envio",
    "bultos": "bultos",
    "kilos": "kilos",
    "observacion": "observacion",
}


def _normalizar_columna(valor):
    texto = str(valor or "").strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = texto.encode("ascii", "ignore").decode("utf-8")
    texto = re.sub(r"\s+", " ", texto)
    return texto


def _texto(valor):
    if pd.isna(valor):
        return ""
    if isinstance(valor, Number) and float(valor).is_integer():
        return str(int(valor)).strip()
    return str(valor).strip()


def _telefono(valor):
    return normalizar_telefono_chile(_texto(valor))


def _numero_entero(valor):
    if pd.isna(valor) or str(valor).strip() == "":
        return None
    try:
        return int(float(valor))
    except (TypeError, ValueError):
        return None


def _es_valor_ejemplo(valor):
    return _normalizar_columna(valor) in {
        "destinatario ejemplo",
        "av. ejemplo 123",
        "av ejemplo 123",
    }


def _obtener_catalogo_comunas(db):
    comunas = db.query(Comuna).order_by(Comuna.c_region.asc(), Comuna.c_nombre.asc()).all()
    por_region = defaultdict(list)
    region_por_comuna = {}

    for comuna in comunas:
        region = comuna.c_region.strip()
        nombre = comuna.c_nombre.strip()
        por_region[region].append(nombre)
        region_por_comuna[_normalizar_columna(nombre)] = region

    return dict(por_region), region_por_comuna


def generar_plantilla_carga_masiva(db):
    """Crea la plantilla oficial que los funcionarios completan para carga masiva."""
    por_region, region_por_comuna = _obtener_catalogo_comunas(db)
    regiones = sorted(por_region)
    comunas_nombres = sorted(
        [nombre for nombres in por_region.values() for nombre in nombres],
        key=_normalizar_columna,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Envios"

    title_fill = PatternFill("solid", fgColor="F3F5F7")
    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="111827", bold=True)
    column_fills = [
        "F8FAFC",
        "F4F7FB",
        "F8FAFC",
        "F4F9F6",
        "F7F9FC",
        "F8FAFC",
        "F4F9F6",
        "F8FAFC",
        "F4F7FB",
        "F8FAFC",
        "F4F9F6",
    ]

    ws.cell(row=1, column=1, value="Datos del remitente").fill = title_fill
    ws.cell(row=1, column=1).font = title_font
    campos_remitente = [
        ("Nombre remitente", "Nombre Apellido"),
        ("Correo remitente", "correo@loreal.com"),
        ("Centro costo", "123456"),
        ("Division", "DPGP"),
    ]
    for row, (etiqueta, ejemplo) in enumerate(campos_remitente, start=2):
        ws.cell(row=row, column=1, value=etiqueta).font = title_font
        ws.cell(row=row, column=2, value=ejemplo)

    ws.cell(row=7, column=1, value="Datos de los envios").fill = title_fill
    ws.cell(row=7, column=1).font = title_font

    for col, titulo in enumerate(COLUMNAS_ENVIOS_PLANTILLA, start=1):
        cell = ws.cell(row=8, column=col, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = max(16, len(titulo) + 4)

    for row in range(9, 309):
        ws.cell(row=row, column=2).number_format = "@"
        ws.cell(row=row, column=6).number_format = "@"
        ws.cell(row=row, column=5, value=f'=IFERROR(VLOOKUP(D{row},Comunas!$A:$B,2,FALSE),"")')
        for col, color in enumerate(column_fills, start=1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=color)

    listas = wb.create_sheet("Listas")
    for row, region in enumerate(regiones, start=1):
        listas.cell(row=row, column=1, value=region)
    for row, division in enumerate(DIVISIONES, start=1):
        listas.cell(row=row, column=2, value=division)
    for row, tipo in enumerate(TIPOS_ENVIO, start=1):
        listas.cell(row=row, column=3, value=tipo)
    for row, comuna in enumerate(comunas_nombres, start=1):
        listas.cell(row=row, column=4, value=comuna)

    comunas_ws = wb.create_sheet("Comunas")
    for row, comuna in enumerate(comunas_nombres, start=1):
        comunas_ws.cell(row=row, column=1, value=comuna)
        comunas_ws.cell(row=row, column=2, value=region_por_comuna[_normalizar_columna(comuna)])

    ejemplo_ws = wb.create_sheet("Ejemplo")
    for col, titulo in enumerate(COLUMNAS_ENVIOS_PLANTILLA, start=1):
        cell = ejemplo_ws.cell(row=1, column=col, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        ejemplo_ws.column_dimensions[cell.column_letter].width = max(16, len(titulo) + 4)
    comuna_ejemplo = comunas_nombres[0] if comunas_nombres else "Comuna"
    ejemplo = [
        "Nombre Destinatario",
        "0",
        "Direccion real 123",
        comuna_ejemplo,
        region_por_comuna.get(_normalizar_columna(comuna_ejemplo), "Region"),
        "912345678",
        "correo.destinatario@empresa.cl",
        "Domicilio",
        1,
        1,
        "Tienda, horario o referencia",
    ]
    for col, valor in enumerate(ejemplo, start=1):
        ejemplo_ws.cell(row=2, column=col, value=valor)

    rango_division = f"Listas!$B$1:$B${len(DIVISIONES)}"
    rango_tipo = f"Listas!$C$1:$C${len(TIPOS_ENVIO)}"
    rango_comunas = f"Listas!$D$1:$D${max(1, len(comunas_nombres))}"
    dv_division = DataValidation(type="list", formula1=f"={rango_division}", allow_blank=False)
    dv_tipo = DataValidation(type="list", formula1=f"={rango_tipo}", allow_blank=False)
    dv_comuna = DataValidation(type="list", formula1=f"={rango_comunas}", allow_blank=False)

    ws.add_data_validation(dv_division)
    ws.add_data_validation(dv_tipo)
    ws.add_data_validation(dv_comuna)
    dv_division.add("B5")
    dv_comuna.add("D9:D308")
    dv_tipo.add("H9:H308")

    ws.freeze_panes = "A9"
    listas.sheet_state = "hidden"
    comunas_ws.sheet_state = "hidden"
    return wb


def _mapear_dataframe(df):
    columnas = {}
    for columna in df.columns:
        normalizada = _normalizar_columna(columna)
        if normalizada in MAPA_COLUMNAS:
            columnas[columna] = MAPA_COLUMNAS[normalizada]

    return df.rename(columns=columnas)


def validar_archivo_carga_masiva(archivo, db):
    """Lee la plantilla, normaliza datos y devuelve filas listas/error para revision web."""
    df = pd.read_excel(archivo, sheet_name="Envios")
    df = _mapear_dataframe(df)
    df = df.dropna(how="all")

    usa_formato_nuevo = "remitente" not in df.columns and "destinatario" not in df.columns
    datos_remitente = None
    if usa_formato_nuevo:
        archivo.seek(0)
        remitente_df = pd.read_excel(
            archivo,
            sheet_name="Envios",
            header=None,
            usecols="A:B",
            nrows=5,
        )
        datos_remitente = {
            "remitente": normalizar_nombre_operativo(
                _texto(remitente_df.iloc[1, 1] if len(remitente_df) > 1 else "")
            ),
            "correo_remitente": _texto(remitente_df.iloc[2, 1] if len(remitente_df) > 2 else ""),
            "centro_costo": _texto(remitente_df.iloc[3, 1] if len(remitente_df) > 3 else ""),
            "division": normalizar_texto_operativo(
                _texto(remitente_df.iloc[4, 1] if len(remitente_df) > 4 else ""),
                upper=True,
            ),
        }

        archivo.seek(0)
        df = pd.read_excel(archivo, sheet_name="Envios", header=7)
        df = _mapear_dataframe(df)
        df = df.dropna(how="all")

    if len(df) > MAX_FILAS_CARGA:
        return {
            "ok": False,
            "errores_archivo": [f"La carga no puede superar {MAX_FILAS_CARGA} filas por archivo"],
            "filas": [],
            "resumen": {},
            "token": None,
        }

    requeridas = [
        "remitente",
        "correo_remitente",
        "centro_costo",
        "division",
        "destinatario",
        "rut_destinatario",
        "direccion",
        "region",
        "comuna",
        "telefono_destinatario",
        "tipo_envio",
        "bultos",
        "kilos",
    ]

    faltantes = [col for col in requeridas if col not in df.columns]
    if datos_remitente:
        faltantes = [
            col for col in faltantes
            if col not in {"remitente", "correo_remitente", "centro_costo", "division"}
        ]
    if faltantes:
        return {
            "ok": False,
            "errores_archivo": ["Faltan columnas obligatorias: " + ", ".join(faltantes)],
            "filas": [],
            "resumen": {},
            "token": None,
        }

    registros = []
    for index, row in df.iterrows():
        remitente = datos_remitente or {}
        registros.append({
            "numero": int(index) + (9 if datos_remitente else 2),
            "remitente": remitente.get("remitente", _texto(row.get("remitente"))),
            "correo_remitente": remitente.get("correo_remitente", _texto(row.get("correo_remitente"))),
            "centro_costo": remitente.get("centro_costo", _texto(row.get("centro_costo"))),
            "division": normalizar_texto_operativo(remitente.get("division", _texto(row.get("division"))), upper=True),
            "destinatario": normalizar_nombre_operativo(_texto(row.get("destinatario"))),
            "rut_destinatario": _texto(row.get("rut_destinatario")),
            "direccion": normalizar_texto_operativo(_texto(row.get("direccion"))),
            "region": normalizar_texto_operativo(_texto(row.get("region"))),
            "comuna": normalizar_texto_operativo(_texto(row.get("comuna"))),
            "telefono_destinatario": _telefono(row.get("telefono_destinatario")),
            "correo_destinatario": _texto(row.get("correo_destinatario")),
            "tipo_envio": _texto(row.get("tipo_envio")).capitalize(),
            "bultos": _numero_entero(row.get("bultos")),
            "kilos": _numero_entero(row.get("kilos")),
            "observacion": normalizar_texto_operativo(_texto(row.get("observacion"))),
        })

    return validar_registros_carga_masiva(registros, db)


def validar_registros_carga_masiva(registros, db):
    por_region, region_por_comuna = _obtener_catalogo_comunas(db)
    regiones_validas = {_normalizar_columna(region): region for region in por_region}
    filas = []
    registros_validos = []

    requeridas = [
        "remitente",
        "correo_remitente",
        "centro_costo",
        "division",
        "destinatario",
        "rut_destinatario",
        "direccion",
        "region",
        "comuna",
        "telefono_destinatario",
        "tipo_envio",
        "bultos",
        "kilos",
    ]

    for index, registro in enumerate(registros):
        data = {
            "remitente": normalizar_nombre_operativo(_texto(registro.get("remitente"))),
            "correo_remitente": _texto(registro.get("correo_remitente")),
            "centro_costo": _texto(registro.get("centro_costo")),
            "division": normalizar_texto_operativo(_texto(registro.get("division")), upper=True),
            "destinatario": normalizar_nombre_operativo(_texto(registro.get("destinatario"))),
            "rut_destinatario": _texto(registro.get("rut_destinatario")),
            "direccion": normalizar_texto_operativo(_texto(registro.get("direccion"))),
            "region": normalizar_texto_operativo(_texto(registro.get("region"))),
            "comuna": normalizar_texto_operativo(_texto(registro.get("comuna"))),
            "telefono_destinatario": _telefono(registro.get("telefono_destinatario")),
            "correo_destinatario": _texto(registro.get("correo_destinatario")),
            "tipo_envio": _texto(registro.get("tipo_envio")).capitalize(),
            "bultos": _numero_entero(registro.get("bultos")),
            "kilos": _numero_entero(registro.get("kilos")),
            "observacion": normalizar_texto_operativo(_texto(registro.get("observacion"))),
        }
        errores = []
        advertencias = []

        comuna_normalizada = _normalizar_columna(data["comuna"])
        region_de_comuna = region_por_comuna.get(comuna_normalizada)
        if region_de_comuna:
            data["region"] = region_de_comuna

        for campo in requeridas:
            if data[campo] in {"", None}:
                errores.append(f"{campo}: obligatorio")

        if data["correo_remitente"] and not email_valido(data["correo_remitente"]):
            errores.append("correo_remitente: formato invalido")

        if _es_valor_ejemplo(data["destinatario"]) or _es_valor_ejemplo(data["direccion"]):
            errores.append("fila de ejemplo: reemplaza destinatario y direccion por datos reales")

        if data["correo_destinatario"] and not email_valido(data["correo_destinatario"]):
            errores.append("correo_destinatario: formato invalido")

        if data["rut_destinatario"] and not rut_operativo_valido(data["rut_destinatario"]):
            errores.append("rut_destinatario: ingresa RUT o 0")

        if data["telefono_destinatario"] and not telefono_chile_valido(data["telefono_destinatario"]):
            errores.append("telefono_destinatario: debe tener 8 o 9 digitos")

        if data["division"] and data["division"] not in DIVISIONES:
            errores.append("division: valor no permitido")

        if data["tipo_envio"] and data["tipo_envio"] not in TIPOS_ENVIO:
            errores.append("tipo_envio: usa Domicilio o Agencia")

        if data["bultos"] is not None and not 1 <= data["bultos"] <= 9999:
            errores.append("bultos: debe estar entre 1 y 9999")

        if data["kilos"] is not None and not 1 <= data["kilos"] <= 9999:
            errores.append("kilos: debe estar entre 1 y 9999")

        region_normalizada = _normalizar_columna(data["region"])
        region_catalogo = regiones_validas.get(region_normalizada)

        if data["region"] and not region_catalogo:
            errores.append("region: no existe en el catalogo")

        if data["comuna"] and not region_de_comuna:
            errores.append("comuna: no existe en el catalogo")

        if region_catalogo and region_de_comuna and region_catalogo != region_de_comuna:
            errores.append("comuna: no corresponde a la region seleccionada")

        if data["tipo_envio"] == "Agencia":
            advertencias.append("Requiere completar codigo de agencia antes de generar lote")

        estado = "error" if errores else ("advertencia" if advertencias else "listo")
        fila = {
            "numero": int(registro.get("numero") or index + 2),
            "data": data,
            "errores": errores,
            "advertencias": advertencias,
            "estado": estado,
        }
        filas.append(fila)

        if not errores:
            registros_validos.append(data)

    token = None
    if registros_validos and not any(fila["errores"] for fila in filas):
        token = guardar_carga_temporal(registros_validos)

    return {
        "ok": not any(fila["errores"] for fila in filas),
        "errores_archivo": [],
        "filas": filas,
        "resumen": {
            "total": len(filas),
            "listos": sum(1 for fila in filas if fila["estado"] == "listo"),
            "advertencias": sum(1 for fila in filas if fila["estado"] == "advertencia"),
            "errores": sum(1 for fila in filas if fila["estado"] == "error"),
        },
        "token": token,
    }


def guardar_carga_temporal(registros):
    TMP_CARGAS_DIR.mkdir(exist_ok=True)
    token = uuid.uuid4().hex
    ruta = TMP_CARGAS_DIR / f"{token}.json"
    ruta.write_text(json.dumps(registros, ensure_ascii=False), encoding="utf-8")
    return token


def leer_carga_temporal(token):
    if not re.fullmatch(r"[a-f0-9]{32}", token or ""):
        return None

    ruta = TMP_CARGAS_DIR / f"{token}.json"
    if not ruta.exists():
        return None

    return json.loads(ruta.read_text(encoding="utf-8"))


def eliminar_carga_temporal(token):
    if not re.fullmatch(r"[a-f0-9]{32}", token or ""):
        return

    ruta = TMP_CARGAS_DIR / f"{token}.json"
    if ruta.exists():
        ruta.unlink()


def construir_envio_desde_carga(data):
    return Envio(
        e_estado="pendiente",
        e_remitente=data["remitente"],
        e_correo_remitente=data["correo_remitente"],
        e_division=data["division"],
        e_centro_costo=data["centro_costo"],
        e_destinatario=data["destinatario"],
        e_rut_destinatario=data["rut_destinatario"],
        e_direccion=data["direccion"],
        e_comuna=data["comuna"],
        e_region=data["region"],
        e_telefono_destinatario=data["telefono_destinatario"],
        e_correo_destinatario=data.get("correo_destinatario", ""),
        e_observacion=data.get("observacion", ""),
        e_tipo_envio=data["tipo_envio"],
        e_codigo_agencia="",
        e_bultos=data["bultos"],
        e_kilos=data["kilos"],
    )
